import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import csv
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl.styles import Color, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from datetime import datetime, timedelta
import time
import os
import openpyxl

user = os.environ.get('USER_WEB_SCRAPPING_PJ')
password = os.environ.get('PASSWORD_WEB_SCRAPPING_PJ')


def first_check():
    if not os.path.isfile('data.xlsx'):
        df = pd.DataFrame(columns=['CHEQUEADO', 'CD', 'Expediente', 'Año', 'Fecha_vencimiento', 'Tipo', 'Descripcion', 'Auto'])
        df.to_excel('data.xlsx', index=False)

    if os.path.exists('error.txt'):
        os.remove('error.txt') 




def prepare_input():
    with open('input.txt', 'r') as f:
        lines = f.readlines()
    values = {}
    for line in lines:
        line = line.strip()
        key, value = line.split(',')
        values[key] = value
    return values




def set_up(u,p):
    browser = webdriver.Chrome('/path/to/chromedriver')
    browser.get('https://scw.pjn.gov.ar/scw/home.seam')
    log_in = browser.find_element('xpath', '//*[@id="nav-header"]/ul/li[2]/a')
    log_in.click()
    time.sleep(1)
    username = browser.find_element('xpath', '/html/body/div[3]/div[2]/div/div[2]/div[1]/div/div/div/form/div[1]/input')
    username.send_keys(u)
    password = browser.find_element('xpath', '//*[@id="password"]')
    password.send_keys(p)
    enter = browser.find_element('xpath','//*[@id="kc-login"]')
    enter.click()
    time.sleep(1)
    query = browser.find_element('xpath','//*[@id="root"]/div/div[1]/div/nav/ul/div[2]')
    query.click()
    time.sleep(1)
    handles = browser.window_handles
    browser.switch_to.window(handles[-1])
    time.sleep(1)
    return browser




def fetch_data(required_values,required_browser):
    browser = required_browser
    values = required_values
    count_errors = 0
    for file, yyear in values.items():
        try:
            set_up_results = browser.find_element('xpath','//*[@id="j_idt83:j_idt85"]/a')
            set_up_results.click()
            time.sleep(1)
            jurisdiction = browser.find_element('xpath', '//*[@id="j_idt83:consultaExpediente:camara"]')
            jurisdiction.send_keys('CIV - Cámara Nacional de Apelaciones en lo Civil')
            number = browser.find_element('xpath', '//*[@id="j_idt83:consultaExpediente:j_idt116:numero"]')
            number.clear()
            number.send_keys(file)
            year = browser.find_element('xpath', '//*[@id="j_idt83:consultaExpediente:j_idt118:anio"]')
            year.clear()
            year.send_keys(yyear)
            query2 = browser.find_element('xpath','//*[@id="j_idt83:consultaExpediente:consultaFiltroSearchButtonSAU"]')
            query2.click()
            details = browser.find_element('xpath','//*[@id="tablaConsultaLista:tablaConsultaForm:j_idt179:dataTable"]/tbody/tr/td[6]/div/a')
            details.click()
            time.sleep(1)
            date = browser.find_element('xpath', '//*[@id="expediente:action-table"]/tbody/tr[1]/td[3]/span[2]').text
            type = browser.find_element('xpath', '//*[@id="expediente:action-table"]/tbody/tr[1]/td[4]/span[2]').text
            description = browser.find_element('xpath', '//*[@id="expediente:action-table"]/tbody/tr[1]/td[5]/span[2]').text
            auto = browser.find_element('xpath', '//*[@id="expediente:j_idt90:detailCover"]')
            new_row = {
                'Expediente': file, 
                'Año': yyear,
                'Fecha_vencimiento':date,
                'Tipo': type,
                'Descripcion': description,
                'Auto': auto.text
            }
            df = pd.DataFrame(new_row, index=[0])
            excel_file = pd.read_excel('data.xlsx',dtype={'Expediente': str,'Año': str,'Fecha_vencimiento': 'str'})
            
            if any((excel_file['Expediente'] == file) & (excel_file['Año'].astype(str).str[:4] == str(yyear))):
                condition = (excel_file['Año'] == yyear) & (excel_file['Expediente'] == file)
                df_with_condition = excel_file[condition]
                chequeado = df_with_condition["CHEQUEADO"].values[0]
                CD = df_with_condition["CD"].values[0]
                df["CHEQUEADO"] = chequeado
                df["CD"] = CD
                excel_file.drop(excel_file[condition].index, inplace=True)        
                excel_file = excel_file.append(df, ignore_index=True)
                excel_file["Fecha_vencimiento"] = pd.to_datetime(excel_file["Fecha_vencimiento"]).dt.strftime("%Y-%m-%d")
                excel_file = excel_file.sort_values(by=['Fecha_vencimiento'], ascending=True)
                excel_file.to_excel('data.xlsx', index=False)
            else:
                excel_file = excel_file.append(df, ignore_index=True)
                excel_file["Fecha_vencimiento"] = pd.to_datetime(excel_file["Fecha_vencimiento"]).dt.strftime("%Y-%m-%d")
                excel_file = excel_file.sort_values(by=['Fecha_vencimiento'], ascending=True)
                excel_file.to_excel('data.xlsx', index=False)
            time.sleep(1)
            back = browser.find_element('xpath','//*[@id="expediente:j_idt78"]/div/a')
            back.click()
            time.sleep(1)
            browser.refresh()
            time.sleep(1)
        except:
            count_errors+=1
            if count_errors==1 and os.path.exists('error.txt'):
                os.remove('error.txt')
            error = "El programa fallo para el expediente {}, año {}. Correrlo de nuevo para esos casos.".format(file,yyear)
            with open("error.txt", "a") as f:
                f.write(error)
                f.write('\n')




def design_table():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook.active
    worksheet.column_dimensions['A'].width = 13
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 10
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 25
    worksheet.column_dimensions['G'].width = 30
    worksheet.column_dimensions['H'].width = 100
    medium_3 = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium3", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    final_excel = pd.read_excel('data.xlsx')
    table = openpyxl.worksheet.table.Table(ref=f"A1:H{len(final_excel)+1}", displayName="Tabla")
    table.tableStyleInfo = medium_3
    worksheet.add_table(table)
    
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        if row[0].value == "OK" or row[0].value == "ok":
            for cell in row:
                cell.font = Font(bold=True)

    workbook.save('data.xlsx')




def time_spent(start):
    start_time = start
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"El tiempo de ejecución fue de: {elapsed_time} segundos")

    


def main():
    start_time = time.time()
    first_check()
    values = prepare_input()
    browser = set_up()
    fetch_data(values,browser)
    design_table()
    time_spent(start_time)


if __name__ == '__main__':
    main()








